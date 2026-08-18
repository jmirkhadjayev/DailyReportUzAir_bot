"""Hisobotlarni Excel (.xlsx) formatiga chiqarish."""
from __future__ import annotations

from datetime import date
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import PageSetupProperties

from config import FOOTER_TEXT, ORG_NAME, chief_name
from i18n import t
from utils import fmt_date, now_str, period_title

COLUMN_KEYS = [
    ("col_num", 5),
    ("col_tabel", 10),
    ("col_date", 12),
    ("col_name", 26),
    ("col_position", 22),
    ("col_done", 55),
    ("col_problems", 35),
    ("col_plans", 35),
    ("col_sent_at", 18),
]

THIN = Side(style="thin", color="B7B7B7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
STRIPE_FILL = PatternFill("solid", fgColor="F2F7FC")
TITLE_FONT = Font(bold=True, size=14, color="1F4E78")
HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
FOOTER_FONT = Font(size=9, italic=True, color="888888")


def _add_signature(ws, row: int, last_col: int, lang: str) -> None:
    """Jadval ostiga imzo joyi: lavozim — chiziq — boshliq F.I.Sh."""
    end_col = min(4, last_col)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=end_col)
    cell = ws.cell(
        row=row,
        column=1,
        value=f"{t(lang, 'doc_sign_label')}    ______________________    {chief_name(lang)}",
    )
    cell.font = Font(size=11, bold=True)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 22

    ws.merge_cells(start_row=row + 1, start_column=1, end_row=row + 1, end_column=end_col)
    hint = ws.cell(row=row + 1, column=1, value=f"{' ' * 28}{t(lang, 'doc_sign_hint')}")
    hint.font = Font(size=9, italic=True, color="888888")
    hint.alignment = Alignment(horizontal="left", vertical="top")


def _add_footer(ws, row: int, last_col: int) -> None:
    """Varaq pastiga imzo qatorini qo'yadi (ekranda va bosmada)."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_col)
    cell = ws.cell(row=row, column=1, value=FOOTER_TEXT)
    cell.font = FOOTER_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.oddFooter.center.text = FOOTER_TEXT
    ws.oddFooter.center.size = 9
    ws.evenFooter.center.text = FOOTER_TEXT
    ws.evenFooter.center.size = 9


def _estimate_height(row, widths: list[int]) -> int:
    """Matn uzunligiga qarab qator balandligini taxminlaydi."""
    lines = 1
    for value, width in zip(row, widths):
        text = str(value or "")
        count = sum(max(1, (len(part) // width) + 1) for part in text.split("\n")) or 1
        lines = max(lines, count)
    return min(15 * min(lines, 20) + 4, 300)


def build_excel(
    rows, date_from: date, date_to: date, lang: str = "uz", subject: str = ""
) -> bytes:
    headers = [(t(lang, key), width) for key, width in COLUMN_KEYS]

    wb = Workbook()
    ws = wb.active
    ws.title = t(lang, "sheet_reports")
    ws.sheet_view.showGridLines = False

    last_col = len(headers)
    widths = [w for _, w in headers]

    # --- sarlavha bloki
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    ws.cell(row=1, column=1, value=ORG_NAME).font = TITLE_FONT
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)
    ws.cell(
        row=2,
        column=1,
        value=f"{t(lang, 'doc_subtitle')}  |  {period_title(date_from, date_to, lang)}",
    )
    ws.cell(row=2, column=1).font = Font(bold=True, size=11)
    ws.cell(row=2, column=1).alignment = Alignment(horizontal="center")

    row = 3
    if subject:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_col)
        cell = ws.cell(row=row, column=1, value=subject)
        cell.font = Font(bold=True, size=11, color="1F4E78")
        cell.alignment = Alignment(horizontal="center")
        row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_col)
    ws.cell(
        row=row,
        column=1,
        value=f"{t(lang, 'doc_generated')}: {now_str()}  |  "
        f"{t(lang, 'doc_total')}: {len(rows)}",
    )
    ws.cell(row=row, column=1).font = Font(size=9, italic=True, color="666666")
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")

    # --- jadval sarlavhasi
    header_row = row + 2
    for col, (title, width) in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=title)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[header_row].height = 30

    # --- ma'lumotlar
    for i, r in enumerate(rows, start=1):
        values = [
            i,
            r["tabel"] or "—",
            fmt_date(r["report_date"]),
            r["full_name"],
            r["position"] or "—",
            r["done"] or "—",
            r["problems"] or "—",
            r["plans"] or "—",
            r["updated_at"],
        ]
        excel_row = header_row + i
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=excel_row, column=col, value=value)
            cell.border = BORDER
            cell.alignment = Alignment(
                horizontal="center" if col in (1, 2, 3, 9) else "left",
                vertical="top",
                wrap_text=col in (4, 5, 6, 7, 8),
            )
            if i % 2 == 0:
                cell.fill = STRIPE_FILL
        ws.row_dimensions[excel_row].height = _estimate_height(values, widths)

    _add_signature(ws, header_row + len(rows) + 2, last_col, lang)
    _add_footer(ws, header_row + len(rows) + 5, last_col)

    if rows:
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(last_col)}{header_row + len(rows)}"
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    ws.print_title_rows = f"{header_row}:{header_row}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)

    _summary_sheet(wb, rows, lang)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _summary_sheet(wb: Workbook, rows, lang: str) -> None:
    ws = wb.create_sheet(t(lang, "sheet_summary"))
    ws.sheet_view.showGridLines = False

    per_employee: dict[str, dict] = {}
    for r in rows:
        item = per_employee.setdefault(
            r["full_name"],
            {"tabel": r["tabel"] or "—", "position": r["position"] or "—", "count": 0, "last": ""},
        )
        item["count"] += 1
        item["last"] = max(item["last"], r["report_date"])

    headers = [
        (t(lang, "col_num"), 5),
        (t(lang, "col_tabel"), 10),
        (t(lang, "col_name"), 30),
        (t(lang, "col_position"), 28),
        (t(lang, "col_report_count"), 16),
        (t(lang, "col_last_report"), 16),
    ]
    for col, (title, width) in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[1].height = 26

    ordered = sorted(per_employee.items(), key=lambda kv: (-kv[1]["count"], kv[0].lower()))
    for i, (name, item) in enumerate(ordered, start=1):
        values = [i, item["tabel"], name, item["position"], item["count"], fmt_date(item["last"])]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=i + 1, column=col, value=value)
            cell.border = BORDER
            cell.alignment = Alignment(
                horizontal="left" if col in (3, 4) else "center", vertical="center"
            )

    total_row = len(ordered) + 2
    ws.cell(row=total_row, column=3, value=t(lang, "total_row")).font = Font(bold=True)
    ws.cell(row=total_row, column=5, value=len(rows)).font = Font(bold=True)
    ws.cell(row=total_row, column=5).alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"
    _add_footer(ws, total_row + 2, len(headers))
